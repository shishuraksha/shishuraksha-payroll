#!/usr/bin/env python3
import openpyxl
from openpyxl import load_workbook
import json

def analyze_payroll_excel(filepath):
    """Analyze the Excel payroll sheet and extract formulas"""
    wb = load_workbook(filepath, data_only=False)
    sheet = wb.active
    
    analysis = {
        "sheet_name": sheet.title,
        "columns": [],
        "formulas": {},
        "attendance_logic": {},
        "calculation_methods": {}
    }
    
    # Get column headers
    headers = {}
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        if cell.value:
            headers[col] = cell.value
            analysis["columns"].append(cell.value)
    
    # Analyze formulas in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 10)):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                col_name = headers.get(cell.column, f"Column {cell.column}")
                if col_name not in analysis["formulas"]:
                    analysis["formulas"][col_name] = []
                analysis["formulas"][col_name].append({
                    "cell": f"{cell.column_letter}{cell.row}",
                    "formula": cell.value
                })
    
    # Look for specific attendance-related formulas
    attendance_columns = []
    for col, header in headers.items():
        if any(keyword in str(header).lower() for keyword in ['attendance', 'present', 'absent', 'working', 'days']):
            attendance_columns.append((col, header))
    
    # Extract attendance calculation logic
    for col, header in attendance_columns:
        formulas = []
        for row in range(2, min(sheet.max_row + 1, 10)):
            cell = sheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append(cell.value)
        if formulas:
            analysis["attendance_logic"][header] = formulas
    
    # Look for salary calculation patterns
    salary_columns = []
    for col, header in headers.items():
        if any(keyword in str(header).lower() for keyword in ['salary', 'basic', 'hra', 'allowance', 'gross', 'net', 'deduction']):
            salary_columns.append((col, header))
    
    # Extract salary calculation methods
    for col, header in salary_columns:
        formulas = []
        for row in range(2, min(sheet.max_row + 1, 10)):
            cell = sheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append(cell.value)
        if formulas:
            analysis["calculation_methods"][header] = formulas[0] if formulas else None
    
    # Print analysis
    print("=== EXCEL PAYROLL SHEET ANALYSIS ===\n")
    print(f"Sheet Name: {analysis['sheet_name']}")
    print(f"Total Columns: {len(analysis['columns'])}")
    print("\nColumn Headers:")
    for i, col in enumerate(analysis['columns'], 1):
        print(f"  {i}. {col}")
    
    print("\n=== ATTENDANCE FORMULAS ===")
    if analysis["attendance_logic"]:
        for col, formulas in analysis["attendance_logic"].items():
            print(f"\n{col}:")
            for formula in formulas[:3]:  # Show first 3 examples
                print(f"  {formula}")
    else:
        print("No attendance formulas found")
    
    print("\n=== SALARY CALCULATION FORMULAS ===")
    if analysis["calculation_methods"]:
        for col, formula in analysis["calculation_methods"].items():
            print(f"\n{col}:")
            print(f"  {formula}")
    else:
        print("No salary calculation formulas found")
    
    # Save detailed analysis
    with open('/home/bhara/payroll/excel_analysis.json', 'w') as f:
        json.dump(analysis, f, indent=2)
    
    return analysis

# Analyze the file
try:
    analyze_payroll_excel('/mnt/c/Users/bhara/Downloads/Payroll_June_25.xlsx')
except Exception as e:
    print(f"Error analyzing Excel file: {e}")
    import traceback
    traceback.print_exc()